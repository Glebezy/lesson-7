import glob
import io
import os
import zipfile
import csv

from pypdf import PdfReader
from openpyxl import load_workbook

files_path = 'files'
resources_path = 'resources'

with zipfile.ZipFile(os.path.join(resources_path, "test.zip"), "w", zipfile.ZIP_DEFLATED) as z:
    file_list = glob.glob(os.path.join(files_path, "*"), recursive=True)
    for file in file_list:
        z.write(file)


def compare_pdf_files(reader1, reader2):
    assert len(reader1.pages) == len(reader2.pages)

    for pageNum in range(len(reader1.pages)):
        text_from_zip = reader1.pages[pageNum].extract_text()
        text_from_file = reader2.pages[pageNum].extract_text()
        assert text_from_zip == text_from_file


def compare_xlsx_files(workbook1, workbook2):
    assert len(workbook1.sheetnames) == len(workbook2.sheetnames)

    for sheet_name in workbook1.sheetnames:
        sheet1 = workbook1[sheet_name]
        sheet2 = workbook2[sheet_name]

        assert sheet1.max_row == sheet2.max_row
        assert sheet1.max_column == sheet2.max_column

        # Сравниваем содержимое ячеек
        for row1, row2 in zip(sheet1.iter_rows(values_only=True), sheet2.iter_rows(values_only=True)):
            for cell1, cell2 in zip(row1, row2):
                assert cell1 == cell2


def compare_csv_files(file1, file2):
    reader1 = csv.reader(file1)
    reader2 = csv.reader(file2)

    assert list(reader1) == list(reader2)


def test_zip_correct_and_content_verification():
    with zipfile.ZipFile(os.path.join(resources_path, "test.zip"), "r") as z:

        for file_name in z.namelist():
            with z.open(file_name) as file:

                if file_name.endswith(".pdf"):
                    reader_from_zip = PdfReader(file_name)
                    reader_from_file = PdfReader(os.path.join(file_name))

                    compare_pdf_files(reader_from_zip, reader_from_file)

                elif file_name.endswith(".xlsx"):
                    workbook_from_zip = load_workbook(file)
                    workbook_from_file = load_workbook(os.path.join(file_name))

                    compare_xlsx_files(workbook_from_zip, workbook_from_file)

                elif file_name.endswith(".csv"):
                    text_stream_from_zip = io.TextIOWrapper(file, encoding='utf-8')

                    with open(os.path.join(file_name), 'r', newline='', encoding='utf-8') as file_from_disk:
                        compare_csv_files(text_stream_from_zip, file_from_disk)
